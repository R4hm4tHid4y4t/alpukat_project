import csv
import io
import os
import aiofiles
from collections import Counter
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, update, delete, and_, or_, desc, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.database import get_db
from app.models.user import User, ProfilPengguna
from app.models.master import Varietas, TingkatKematangan, ModelCnn
from app.models.deteksi import HasilDeteksi, RiwayatDeteksi
from app.dependencies.auth import get_current_admin
from app.schemas.admin import VarietasCRUD, KematanganCRUD, FlagRequest
from app.services.tflite_service import TFLiteInferenceService
from app.utils.response import success_response, paginated_response
from app.config import get_settings

router = APIRouter()
settings = get_settings()


# ── UC-11: DASHBOARD ADMIN ────────────────────────────────
@router.get("/dashboard")
async def dashboard(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    now = datetime.now(timezone.utc)

    total_pengguna = (await db.execute(
        select(func.count()).select_from(User).where(User.role == "pengguna")
    )).scalar()

    total_deteksi = (await db.execute(
        select(func.count()).select_from(HasilDeteksi)
    )).scalar()

    deteksi_hari_ini = (await db.execute(
        select(func.count()).select_from(HasilDeteksi)
        .where(func.date(HasilDeteksi.created_at) == func.curdate())
    )).scalar()

    deteksi_bulan_ini = (await db.execute(
        select(func.count()).select_from(HasilDeteksi)
        .where(
            func.month(HasilDeteksi.created_at) == now.month,
            func.year(HasilDeteksi.created_at) == now.year,
        )
    )).scalar()

    total_flagged = (await db.execute(
        select(func.count()).select_from(HasilDeteksi)
        .where(HasilDeteksi.status_flag == "perlu_ditinjau")
    )).scalar()

    # Akurasi dari model aktif (prioritaskan varietas)
    model_aktif = (await db.execute(
        select(ModelCnn).where(ModelCnn.status_aktif == 1).order_by(ModelCnn.id).limit(1)
    )).scalar_one_or_none()

    # Tren 7 hari terakhir
    tren_q = await db.execute(
        select(
            func.date(HasilDeteksi.created_at).label("tanggal"),
            func.count().label("jumlah"),
        )
        .where(
            HasilDeteksi.created_at >= func.date_sub(func.curdate(), text("interval 7 day"))
        )
        .group_by(func.date(HasilDeteksi.created_at))
        .order_by(func.date(HasilDeteksi.created_at))
    )
    tren_mingguan = [
        {"tanggal": str(r.tanggal), "jumlah": r.jumlah}
        for r in tren_q.all()
    ]

    # Distribusi varietas
    var_q = await db.execute(
        select(Varietas.nama_varietas, func.count(HasilDeteksi.id).label("jumlah"))
        .join(HasilDeteksi, HasilDeteksi.varietas_id == Varietas.id)
        .group_by(Varietas.id)
    )
    var_rows = var_q.all()
    total_var = sum(r.jumlah for r in var_rows)
    distribusi_varietas = [
        {"nama": r.nama_varietas, "jumlah": r.jumlah,
         "persentase": round(r.jumlah / total_var * 100, 1) if total_var else 0}
        for r in var_rows
    ]

    # Distribusi kematangan
    kem_q = await db.execute(
        select(TingkatKematangan.label_kematangan, func.count(HasilDeteksi.id).label("jumlah"))
        .join(HasilDeteksi, HasilDeteksi.kematangan_id == TingkatKematangan.id)
        .group_by(TingkatKematangan.id)
    )
    kem_rows = kem_q.all()
    total_kem = sum(r.jumlah for r in kem_rows)
    distribusi_kematangan = [
        {"label": r.label_kematangan, "jumlah": r.jumlah,
         "persentase": round(r.jumlah / total_kem * 100, 1) if total_kem else 0}
        for r in kem_rows
    ]

    return success_response(data={
        "total_pengguna": total_pengguna,
        "total_deteksi": total_deteksi,
        "deteksi_hari_ini": deteksi_hari_ini,
        "deteksi_bulan_ini": deteksi_bulan_ini,
        "total_flagged": total_flagged,
        "akurasi_model": float(model_aktif.akurasi) if model_aktif and model_aktif.akurasi else None,
        "model_versi": model_aktif.versi if model_aktif else None,
        "tren_mingguan": tren_mingguan,
        "distribusi_varietas": distribusi_varietas,
        "distribusi_kematangan": distribusi_kematangan,
    })


# ── UC-14: MANAJEMEN PENGGUNA ─────────────────────────────
@router.get("/users")
async def list_users(
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    search: str = Query(None),
    role: str = Query(None),
    status: int = Query(None),
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    conditions = []
    if search:
        conditions.append(or_(
            User.nama.ilike(f"%{search}%"),
            User.email.ilike(f"%{search}%"),
        ))
    if role:
        conditions.append(User.role == role)
    if status is not None:
        conditions.append(User.status_verifikasi == status)

    total = (await db.execute(
        select(func.count()).select_from(User).where(and_(*conditions) if conditions else True)
    )).scalar()

    result = await db.execute(
        select(User).where(and_(*conditions) if conditions else True)
        .order_by(desc(User.created_at))
        .limit(per_page).offset((page - 1) * per_page)
    )
    users = result.scalars().all()

    items = []
    for u in users:
        count_q = await db.execute(
            select(func.count()).select_from(HasilDeteksi).where(HasilDeteksi.user_id == u.id)
        )
        items.append({
            **u.to_dict(),
            "total_deteksi": count_q.scalar(),
        })

    return paginated_response(items=items, total=total, page=page, per_page=per_page)


@router.put("/users/{user_id}/toggle-role")
async def toggle_role(
    user_id: int,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    if user_id == admin.id:
        raise HTTPException(status_code=400, detail="Tidak dapat mengubah role diri sendiri")

    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")

    user.role = "admin" if user.role == "pengguna" else "pengguna"
    await db.commit()
    return success_response(
        data={"id": user.id, "role": user.role},
        message=f"Role berhasil diubah menjadi {user.role}",
    )


@router.put("/users/{user_id}/toggle-status")
async def toggle_status(
    user_id: int,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    if user_id == admin.id:
        raise HTTPException(status_code=400, detail="Tidak dapat menonaktifkan akun sendiri")

    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")

    user.status_verifikasi = 0 if user.status_verifikasi == 1 else 1
    await db.commit()
    status_label = "aktif" if user.status_verifikasi == 1 else "nonaktif"
    return success_response(
        data={"id": user.id, "status_verifikasi": user.status_verifikasi},
        message=f"Akun berhasil di{status_label}kan",
    )


# ── UC-12: CRUD VARIETAS ──────────────────────────────────
@router.get("/varietas")
async def list_varietas(admin: User = Depends(get_current_admin), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Varietas).order_by(Varietas.id))
    varietas_list = result.scalars().all()
    items = []
    for v in varietas_list:
        count_q = await db.execute(
            select(func.count()).select_from(HasilDeteksi).where(HasilDeteksi.varietas_id == v.id)
        )
        items.append({**v.to_dict(), "total_deteksi": count_q.scalar()})
    return success_response(data=items)


@router.post("/varietas", status_code=201)
async def create_varietas(
    body: VarietasCRUD,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    exists = (await db.execute(
        select(Varietas).where(func.lower(Varietas.nama_varietas) == body.nama_varietas.lower())
    )).scalar_one_or_none()
    if exists:
        raise HTTPException(status_code=400, detail="Nama varietas sudah ada")

    v = Varietas(**body.model_dump())
    db.add(v)
    await db.commit()
    return success_response(data=v.to_dict(), message="Varietas berhasil ditambahkan", status_code=201)


@router.put("/varietas/{varietas_id}")
async def update_varietas(
    varietas_id: int, body: VarietasCRUD,
    admin: User = Depends(get_current_admin), db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Varietas).where(Varietas.id == varietas_id))
    v = result.scalar_one_or_none()
    if not v:
        raise HTTPException(status_code=404, detail="Varietas tidak ditemukan")
    for key, val in body.model_dump(exclude_none=True).items():
        setattr(v, key, val)
    await db.commit()
    return success_response(data=v.to_dict(), message="Varietas berhasil diperbarui")


@router.delete("/varietas/{varietas_id}")
async def delete_varietas(
    varietas_id: int,
    admin: User = Depends(get_current_admin), db: AsyncSession = Depends(get_db),
):
    count_q = await db.execute(
        select(func.count()).select_from(HasilDeteksi).where(HasilDeteksi.varietas_id == varietas_id)
    )
    if count_q.scalar() > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Varietas masih digunakan dalam {count_q.scalar()} riwayat deteksi"
        )
    result = await db.execute(select(Varietas).where(Varietas.id == varietas_id))
    v = result.scalar_one_or_none()
    if not v:
        raise HTTPException(status_code=404, detail="Varietas tidak ditemukan")
    await db.delete(v)
    await db.commit()
    return success_response(message="Varietas berhasil dihapus")


# ── UC-13: CRUD KEMATANGAN ────────────────────────────────
@router.get("/kematangan")
async def list_kematangan(admin: User = Depends(get_current_admin), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(TingkatKematangan).order_by(TingkatKematangan.id))
    items = result.scalars().all()
    data = []
    for k in items:
        count_q = await db.execute(
            select(func.count()).select_from(HasilDeteksi).where(HasilDeteksi.kematangan_id == k.id)
        )
        data.append({**k.to_dict(), "total_deteksi": count_q.scalar()})
    return success_response(data=data)


@router.post("/kematangan", status_code=201)
async def create_kematangan(
    body: KematanganCRUD,
    admin: User = Depends(get_current_admin), db: AsyncSession = Depends(get_db),
):
    k = TingkatKematangan(**body.model_dump())
    db.add(k)
    await db.commit()
    return success_response(data=k.to_dict(), message="Tingkat kematangan berhasil ditambahkan", status_code=201)


@router.put("/kematangan/{kematangan_id}")
async def update_kematangan(
    kematangan_id: int, body: KematanganCRUD,
    admin: User = Depends(get_current_admin), db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(TingkatKematangan).where(TingkatKematangan.id == kematangan_id))
    k = result.scalar_one_or_none()
    if not k:
        raise HTTPException(status_code=404, detail="Tingkat kematangan tidak ditemukan")
    for key, val in body.model_dump(exclude_none=True).items():
        setattr(k, key, val)
    await db.commit()
    return success_response(data=k.to_dict(), message="Tingkat kematangan berhasil diperbarui")


@router.delete("/kematangan/{kematangan_id}")
async def delete_kematangan(
    kematangan_id: int,
    admin: User = Depends(get_current_admin), db: AsyncSession = Depends(get_db),
):
    count_q = await db.execute(
        select(func.count()).select_from(HasilDeteksi).where(HasilDeteksi.kematangan_id == kematangan_id)
    )
    if count_q.scalar() > 0:
        raise HTTPException(status_code=400, detail="Tingkat kematangan masih digunakan dalam riwayat deteksi")
    result = await db.execute(select(TingkatKematangan).where(TingkatKematangan.id == kematangan_id))
    k = result.scalar_one_or_none()
    if not k:
        raise HTTPException(status_code=404, detail="Tingkat kematangan tidak ditemukan")
    await db.delete(k)
    await db.commit()
    return success_response(message="Tingkat kematangan berhasil dihapus")


# ── UC-15: KELOLA MODEL CNN ───────────────────────────────
# ── UC-15: EVALUASI MODEL CNN ─────────────────────────────
@router.get("/model/evaluasi")
async def get_evaluasi_model(
    admin: User = Depends(get_current_admin), 
    db: AsyncSession = Depends(get_db)
):
    """
    Endpoint untuk mengambil metrik evaluasi dari model yang aktif saat ini
    (opsional jika ingin disambungkan kembali dengan data dinamis UI di kemudian hari).
    """
    result = await db.execute(select(ModelCnn).where(ModelCnn.status_aktif == 1))
    models = result.scalars().all()
    
    return success_response(
        data=[m.to_dict() for m in models],
        message="Data evaluasi model berhasil diambil"
    )

# ── UC-16, UC-17: RIWAYAT DETEKSI GLOBAL ─────────────────
@router.get("/deteksi")
async def list_deteksi_global(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    varietas_id: int = Query(None),
    kematangan_id: int = Query(None),
    flag: str = Query(None),
    search: str = Query(None),
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    conditions = []
    if varietas_id:
        conditions.append(HasilDeteksi.varietas_id == varietas_id)
    if kematangan_id:
        conditions.append(HasilDeteksi.kematangan_id == kematangan_id)
    if flag:
        conditions.append(HasilDeteksi.status_flag == flag)
    if search:
        conditions.append(or_(
            User.nama.ilike(f"%{search}%"),
            User.email.ilike(f"%{search}%"),
        ))

    base_query = (
        select(HasilDeteksi, User, Varietas, TingkatKematangan)
        .join(User, HasilDeteksi.user_id == User.id)
        .outerjoin(Varietas, HasilDeteksi.varietas_id == Varietas.id)
        .outerjoin(TingkatKematangan, HasilDeteksi.kematangan_id == TingkatKematangan.id)
        .where(and_(*conditions) if conditions else True)
    )

    total = (await db.execute(
        select(func.count()).select_from(base_query.subquery())
    )).scalar()

    result = await db.execute(
        base_query.order_by(desc(HasilDeteksi.created_at))
        .limit(per_page).offset((page - 1) * per_page)
    )

    items = []
    for hasil, user, varietas, kematangan in result.all():
        items.append({
            "id": hasil.id,
            "pengguna": {"id": user.id, "nama": user.nama, "email": user.email},
            "varietas_nama": varietas.nama_varietas if varietas else None,
            "kematangan_label": kematangan.label_kematangan if kematangan else None,
            "confidence_varietas": float(hasil.confidence_varietas) if hasil.confidence_varietas else None,
            "confidence_kematangan": float(hasil.confidence_kematangan) if hasil.confidence_kematangan else None,
            "status_flag": hasil.status_flag,
            "catatan_flag": hasil.catatan_flag,
            "created_at": hasil.created_at.isoformat() if hasil.created_at else None,
        })

    return paginated_response(items=items, total=total, page=page, per_page=per_page)


@router.put("/deteksi/{hasil_id}/flag")
async def flag_deteksi(
    hasil_id: int,
    body: FlagRequest,
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(HasilDeteksi).where(HasilDeteksi.id == hasil_id))
    hasil = result.scalar_one_or_none()
    if not hasil:
        raise HTTPException(status_code=404, detail="Hasil deteksi tidak ditemukan")

    hasil.status_flag = body.status_flag
    hasil.catatan_flag = body.catatan_flag

    db.add(RiwayatDeteksi(
        user_id=admin.id,
        hasil_id=hasil_id,
        aksi="ditandai_admin",
    ))
    await db.commit()

    return success_response(
        data={"id": hasil_id, "status_flag": body.status_flag},
        message="Status flag berhasil diperbarui",
    )


@router.get("/deteksi/export")
async def export_deteksi(
    admin: User = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(HasilDeteksi, User, Varietas, TingkatKematangan)
        .join(User, HasilDeteksi.user_id == User.id)
        .outerjoin(Varietas, HasilDeteksi.varietas_id == Varietas.id)
        .outerjoin(TingkatKematangan, HasilDeteksi.kematangan_id == TingkatKematangan.id)
        .order_by(desc(HasilDeteksi.created_at))
    )
    rows = result.all()

    # ── Susun data mentah ────────────────────────────────────────────────────
    data_rows = []
    for hasil, user, varietas, kematangan in rows:
        data_rows.append((
            hasil.id,
            user.nama,
            user.email,
            varietas.nama_varietas if varietas else "-",
            kematangan.label_kematangan if kematangan else "-",
            float(hasil.confidence_varietas) if hasil.confidence_varietas else 0.0,
            float(hasil.confidence_kematangan) if hasil.confidence_kematangan else 0.0,
            hasil.status_flag or "-",
            hasil.created_at if hasil.created_at else None,
        ))

    # ── Konstanta gaya ────────────────────────────────────────────────────────
    C_TITLE    = "FF1F2937"
    C_DOCNUM   = "FF15803D"
    C_BODY     = "FF374151"
    C_GREY     = "FF6B7280"
    FILL_HDR   = PatternFill("solid", fgColor="FFE5E7EB")
    FILL_TOTAL = PatternFill("solid", fgColor="FFF3F4F6")
    thin       = Side(style="thin", color="FFD1D5DB")
    bdr        = Border(left=thin, right=thin, top=thin, bottom=thin)
    FONT       = "Arial"

    COLS   = ["B", "C", "D", "E", "F", "G", "H", "I", "J"]
    WIDTHS = [7, 20, 28, 12, 16, 14, 15, 16, 19]
    LAST_COL = COLS[-1]

    STATUS_STYLE = {
        "perlu_ditinjau": ("Perlu Ditinjau", "FFB45309", "FFFEF3C7"),
        "valid":          ("Valid",          "FF15803D", "FFDCFCE7"),
        "salah":          ("Salah",          "FFB91C1C", "FFFEE2E2"),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Riwayat Deteksi"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2.5
    for letter, w in zip(COLS, WIDTHS):
        ws.column_dimensions[letter].width = w

    def merge_set(row, value, bold=False, size=10, color=C_BODY, align="center", start_col="B"):
        rng = f"{start_col}{row}:{LAST_COL}{row}"
        ws.merge_cells(rng)
        c = ws[f"{start_col}{row}"]
        c.value     = value
        c.font      = Font(name=FONT, bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align, vertical="center")
        return c

    # ── BLOK JUDUL ───────────────────────────────────────────────────────────
    now    = datetime.now()
    month  = now.strftime("%m")
    year   = now.strftime("%Y")
    doc_no = f"RPD/{month}/{year}/{len(data_rows):03d}"
    hari   = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"][now.weekday()]

    merge_set(1, "LAPORAN RIWAYAT DETEKSI", bold=True, size=17, color=C_TITLE)
    merge_set(2, "AvoScan  |  Sistem Klasifikasi & Kematangan Alpukat", size=10, color=C_GREY)
    ws.row_dimensions[3].height = 6  # spacer

    merge_set(4, f"No. Dokumen  :  {doc_no}", bold=True, size=10, color=C_DOCNUM, align="left")
    merge_set(5, f"Periode          :  Seluruh Data (s.d. {now.strftime('%d/%m/%Y')})", size=10, color=C_BODY, align="left")
    merge_set(6, f"Dicetak Pada   :  {hari}, {now.strftime('%d/%m/%Y')} pukul {now.strftime('%H.%M')} WIB", size=10, color=C_BODY, align="left")

    ws.row_dimensions[7].height = 4
    for letter in COLS:
        ws[f"{letter}7"].border = Border(bottom=Side(style="medium", color="FF15803D"))

    # ── HEADER TABEL ─────────────────────────────────────────────────────────
    HEADER_ROW = 9
    headers = ["ID", "Pengguna", "Email", "Varietas", "Kematangan",
               "Conf.\nVarietas", "Conf.\nKematangan", "Status", "Tanggal Deteksi"]
    for letter, hdr in zip(COLS, headers):
        c = ws[f"{letter}{HEADER_ROW}"]
        c.value     = hdr
        c.font      = Font(name=FONT, bold=True, size=9, color=C_TITLE)
        c.fill      = FILL_HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = bdr
    ws.row_dimensions[HEADER_ROW].height = 30

    # ── BARIS DATA ───────────────────────────────────────────────────────────
    DATA_START = HEADER_ROW + 1
    for r_i, (id_, nama, email, varietas, kematangan, cv, ck, flag, tgl_det) in \
            enumerate(data_rows, start=DATA_START):
        label, fcolor, fillcolor = STATUS_STYLE.get(flag, (flag, C_BODY, None))
        tgl_str = tgl_det.strftime("%d/%m/%Y %H:%M") if tgl_det else "-"
        vals_aligns = [
            (id_, "center", None),
            (nama, "left", None),
            (email, "left", None),
            (varietas, "center", None),
            (kematangan, "center", None),
            (cv / 100, "center", None),
            (ck / 100, "center", None),
            (label, "center", "STATUS"),
            (tgl_str, "center", None),
        ]
        for letter, (val, align, kind) in zip(COLS, vals_aligns):
            c = ws[f"{letter}{r_i}"]
            c.value     = val
            c.border    = bdr
            c.alignment = Alignment(horizontal=align, vertical="center")
            if kind == "STATUS":
                c.font = Font(name=FONT, bold=True, size=8.5, color=fcolor)
                if fillcolor:
                    c.fill = PatternFill("solid", fgColor=fillcolor)
            else:
                c.font = Font(name=FONT, size=9, color=C_BODY)
        ws[f"G{r_i}"].number_format = '0.00"%"'
        ws[f"H{r_i}"].number_format = '0.00"%"'
        ws.row_dimensions[r_i].height = 20

    # ── BARIS TOTAL ──────────────────────────────────────────────────────────
    TOTAL_ROW = DATA_START + len(data_rows)
    ws.merge_cells(f"B{TOTAL_ROW}:E{TOTAL_ROW}")
    c = ws[f"B{TOTAL_ROW}"]
    c.value     = f"TOTAL  :  {len(data_rows)} REKAMAN"
    c.font      = Font(name=FONT, bold=True, size=9.5, color=C_TITLE)
    c.fill      = FILL_TOTAL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = bdr

    avg_cv = sum(r[5] for r in data_rows) / len(data_rows) if data_rows else 0
    avg_ck = sum(r[6] for r in data_rows) / len(data_rows) if data_rows else 0

    ws[f"F{TOTAL_ROW}"].value     = "Rata-rata"
    ws[f"F{TOTAL_ROW}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"F{TOTAL_ROW}"].font      = Font(name=FONT, bold=True, size=8.5, color=C_TITLE)
    ws[f"G{TOTAL_ROW}"].value     = avg_cv / 100
    ws[f"G{TOTAL_ROW}"].number_format = '0.00"%"'
    ws[f"H{TOTAL_ROW}"].value     = avg_ck / 100
    ws[f"H{TOTAL_ROW}"].number_format = '0.00"%"'

    for letter in ["F", "G", "H", "I", "J"]:
        ws[f"{letter}{TOTAL_ROW}"].fill   = FILL_TOTAL
        ws[f"{letter}{TOTAL_ROW}"].border = bdr
        if letter in ("G", "H"):
            ws[f"{letter}{TOTAL_ROW}"].font      = Font(name=FONT, bold=True, size=9, color=C_TITLE)
            ws[f"{letter}{TOTAL_ROW}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[TOTAL_ROW].height = 22

    # ── STATISTIK DISTRIBUSI ─────────────────────────────────────────────────
    STAT_ROW  = TOTAL_ROW + 2
    var_count = Counter(r[3] for r in data_rows)
    mat_count = Counter(r[4] for r in data_rows)
    var_text  = "   ".join(f"{k} ({v})" for k, v in var_count.items()) or "-"
    mat_text  = "   ".join(f"{k} ({v})" for k, v in mat_count.items()) or "-"

    ws.merge_cells(f"B{STAT_ROW}:{LAST_COL}{STAT_ROW}")
    ws[f"B{STAT_ROW}"].value = f"Distribusi Varietas:  {var_text}"
    ws[f"B{STAT_ROW}"].font  = Font(name=FONT, size=8.5, color=C_GREY, italic=True)

    ws.merge_cells(f"B{STAT_ROW + 1}:{LAST_COL}{STAT_ROW + 1}")
    ws[f"B{STAT_ROW + 1}"].value = f"Distribusi Kematangan:  {mat_text}"
    ws[f"B{STAT_ROW + 1}"].font  = Font(name=FONT, size=8.5, color=C_GREY, italic=True)

    # ── CATATAN & TANDA TANGAN ───────────────────────────────────────────────
    NOTE_ROW = STAT_ROW + 3
    ws[f"B{NOTE_ROW}"].value = "Catatan:"
    ws[f"B{NOTE_ROW}"].font  = Font(name=FONT, size=8.5, bold=True, color=C_GREY)
    for i, note in enumerate([
        "1. Data dihasilkan otomatis oleh sistem AvoScan berdasarkan hasil inferensi model CNN.",
        "2. Status \"Perlu Ditinjau\" menandakan hasil deteksi perlu diverifikasi kembali secara manual.",
    ], start=1):
        ws.merge_cells(f"B{NOTE_ROW + i}:F{NOTE_ROW + i}")
        ws[f"B{NOTE_ROW + i}"].value = note
        ws[f"B{NOTE_ROW + i}"].font  = Font(name=FONT, size=8, color=C_GREY)

    ws.merge_cells(f"G{NOTE_ROW}:{LAST_COL}{NOTE_ROW}")
    ws[f"G{NOTE_ROW}"].value     = f"Padang, {now.strftime('%d %B %Y')}"
    ws[f"G{NOTE_ROW}"].font      = Font(name=FONT, size=9, color=C_BODY)
    ws[f"G{NOTE_ROW}"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"G{NOTE_ROW + 1}:{LAST_COL}{NOTE_ROW + 1}")
    ws[f"G{NOTE_ROW + 1}"].value     = "Mengetahui,"
    ws[f"G{NOTE_ROW + 1}"].font      = Font(name=FONT, size=9, color=C_BODY)
    ws[f"G{NOTE_ROW + 1}"].alignment = Alignment(horizontal="center")

    SIGN_ROW = NOTE_ROW + 5
    top_border = Border(top=Side(style="thin", color="FF9CA3AF"))
    ws.merge_cells(f"G{SIGN_ROW}:{LAST_COL}{SIGN_ROW}")
    ws[f"G{SIGN_ROW}"].value     = admin.nama or "Admin Sistem"
    ws[f"G{SIGN_ROW}"].font      = Font(name=FONT, bold=True, size=10, color=C_TITLE)
    ws[f"G{SIGN_ROW}"].alignment = Alignment(horizontal="center")
    for letter in COLS[COLS.index("G"):]:
        ws[f"{letter}{SIGN_ROW}"].border = top_border

    ws.merge_cells(f"G{SIGN_ROW + 1}:{LAST_COL}{SIGN_ROW + 1}")
    ws[f"G{SIGN_ROW + 1}"].value     = "Administrator"
    ws[f"G{SIGN_ROW + 1}"].font      = Font(name=FONT, size=8.5, color=C_GREY)
    ws[f"G{SIGN_ROW + 1}"].alignment = Alignment(horizontal="center")

    # Header tabel tetap terlihat saat scroll ke bawah
    ws.freeze_panes = f"B{DATA_START}"

    # Pengaturan cetak: landscape, muat ke lebar 1 halaman
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    # ── Simpan ke bytes & kembalikan ─────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today    = now.strftime("%Y%m%d")
    filename = f"laporan_riwayat_deteksi_{today}.xlsx"
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )